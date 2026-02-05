"""
XLSX spreadsheet processor for Excel files.
"""

import logging
from pathlib import Path
from typing import Generator

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as XLImage
from copy import copy
import io

from .base import BaseFileProcessor, ContentChunk, ContentType

logger = logging.getLogger("OfficeTranslator.XlsxProcessor")


class XlsxProcessor(BaseFileProcessor):
    """Processor for Microsoft Excel (.xlsx) spreadsheets."""
    
    SUPPORTED_EXTENSIONS = ["xlsx"]
    
    def __init__(self):
        super().__init__()
        self._document: Workbook = None
        self._cell_map: dict[str, Cell] = {}  # id -> cell reference
        self._image_map: dict[str, tuple] = {} # id -> (sheet_name, anchor)
    
    def load(self, file_path: str | Path) -> None:
        """Load an XLSX file."""
        path = Path(file_path)
        
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        
        if path.suffix.lower() != ".xlsx":
            raise ValueError(f"Not an XLSX file: {path}")
        
        try:
            self._document = load_workbook(path)
            self._file_path = path
            self._cell_map.clear()
            self._image_map.clear()
            logger.info(f"Loaded XLSX: {path.name} ({len(self._document.sheetnames)} sheets)")
        except Exception as e:
            raise ValueError(f"Invalid XLSX file: {e}")
    
    def extract_content_generator(self) -> Generator[ContentChunk, None, None]:
        """Extract text content and images from all sheets."""
        self._validate_loaded()
        
        chunk_count = 0
        
        for sheet_name in self._document.sheetnames:
            sheet = self._document[sheet_name]
            
            # Process Rows
            for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value and isinstance(cell.value, str):
                        cell_text = str(cell.value).strip()
                        if cell_text:
                            chunk_id = f"{sheet_name}_{cell.coordinate}"
                            self._cell_map[chunk_id] = cell
                            
                            yield ContentChunk(
                                id=chunk_id,
                                content_type=ContentType.TABLE_CELL,
                                text=cell_text,
                                location=f"Sheet '{sheet_name}', Cell {cell.coordinate}",
                                metadata={
                                    "sheet": sheet_name,
                                    "row": row_idx,
                                    "col": col_idx,
                                    "coordinate": cell.coordinate,
                                    "has_formula": cell_text.startswith("="),
                                }
                            )
                            chunk_count += 1
            
            # Process Images
            if hasattr(sheet, '_images'):
                for idx, img in enumerate(sheet._images):
                    try:
                        # Try to get image bytes
                        img_bytes = None
                        if hasattr(img, '_data'):
                            img_bytes = img._data()
                        elif hasattr(img, 'ref') and hasattr(img.ref, 'read'):
                            img.ref.seek(0)
                            img_bytes = img.ref.read()
                        
                        if img_bytes:
                            chunk_id = f"img_{sheet_name}_{idx}"
                            # Store anchor to place translated image
                            self._image_map[chunk_id] = (sheet_name, img.anchor)
                            
                            yield ContentChunk(
                                id=chunk_id,
                                content_type=ContentType.IMAGE,
                                image_data=img_bytes,
                                location=f"Image in {sheet_name}",
                            )
                            chunk_count += 1
                    except Exception as e:
                        logger.debug(f"Failed to extract image from {sheet_name}: {e}")
        
        logger.info(f"Extracted {chunk_count} chunks from XLSX")
    
    def apply_translation(
        self,
        chunk_id: str,
        translated_content: str | bytes,
    ) -> None:
        """Apply translation to a cell or image."""
        self._validate_loaded()
        
        # Handle Image
        if chunk_id in self._image_map:
            if isinstance(translated_content, bytes):
                try:
                    sheet_name, anchor = self._image_map[chunk_id]
                    sheet = self._document[sheet_name]
                    
                    new_img = XLImage(io.BytesIO(translated_content))
                    new_img.anchor = anchor
                    sheet.add_image(new_img)
                    logger.debug(f"Applied translated image to {chunk_id}")
                except Exception as e:
                    logger.warning(f"Failed to apply image translation {chunk_id}: {e}")
            return

        if chunk_id not in self._cell_map:
            logger.warning(f"Chunk not found: {chunk_id}")
            return
        
        cell = self._cell_map[chunk_id]
        
        # Don't translate formulas
        if cell.value and str(cell.value).startswith("="):
            logger.debug(f"Skipping formula cell: {chunk_id}")
            return
        
        cell.value = str(translated_content)
        
        # Ensure text wraps to avoid layout issues with longer translations
        # We try to preserve existing alignment properties
        if cell.alignment:
            new_align = copy(cell.alignment)
            new_align.wrap_text = True
            cell.alignment = new_align
        else:
            cell.alignment = Alignment(wrap_text=True)
            
        logger.debug(f"Applied translation to {chunk_id}")
    
    def save(self, output_path: str | Path) -> None:
        """Save the modified workbook."""
        self._validate_loaded()
        
        path = Path(output_path)
        self._document.save(path)
        logger.info(f"Saved XLSX: {path.name}")
    
    def get_total_chunks(self) -> int:
        """Efficiently count text cells."""
        self._validate_loaded()
        
        count = 0
        for sheet_name in self._document.sheetnames:
            sheet = self._document[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if not str(cell.value).startswith("="):
                            count += 1
        return count
